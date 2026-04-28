"""Business logic service for Open Data Analyzer project."""

from __future__ import annotations

import asyncio
import os
import tempfile
import uuid
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from typing import Any, Dict, List

import logging

from sqlalchemy.ext.asyncio import AsyncSession

from app.database import async_session_factory
from app.projects.open_data_analyzer.core import (
    GEMINI_BASE_URL,
    GEMINI_MODEL,
    STANDARD_CATEGORIES,
    SimpleTable,
    extract_tables_from_excel,
    normalize_table_key,
    run_stage1,
)
from app.projects.open_data_analyzer.reasons import REASONS
from app.projects.open_data_analyzer.schemas import AnalyzerStats
from app.shared.services.execution_recorder import (
    ExecutionRecorder,
    serialize_execution,
)

logger = logging.getLogger(__name__)

PROJECT_SLUG = "open_data_analyzer"
FIXED_SLEEP_SECONDS = 1.0

SESSIONS: Dict[str, Dict[str, Any]] = {}
# In-memory progress tracker: execution_id → {done, total, status, error?}
_PROGRESS: Dict[str, Dict[str, Any]] = {}


def _save_upload_bytes(data: bytes, suffix: str) -> str:
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    handle.write(data)
    handle.flush()
    handle.close()
    return handle.name


def _snap_to_standard(label: str, threshold: float = 0.82) -> str:
    best_ratio, best_cat = 0.0, label
    for cat in STANDARD_CATEGORIES:
        r = SequenceMatcher(None, label, cat).ratio()
        if r > best_ratio:
            best_ratio, best_cat = r, cat
    return best_cat if best_ratio >= threshold else label


def _consolidate_major_areas(stage1: Dict[str, Dict[str, Any]]) -> None:
    """In-place: snap major_area to nearest standard, then merge near-duplicates."""
    for s1 in stage1.values():
        ma = s1.get("major_area", "")
        if ma:
            s1["major_area"] = _snap_to_standard(ma)

    freq = Counter(s1.get("major_area", "") for s1 in stage1.values() if s1.get("major_area"))
    label_map: Dict[str, str] = {}
    for label in freq:
        if label in STANDARD_CATEGORIES:
            label_map[label] = label
            continue
        best_ratio, best_target = 0.0, label
        for other in freq:
            if other == label:
                continue
            r = SequenceMatcher(None, label, other).ratio()
            if r > best_ratio:
                best_ratio, best_target = r, other
        label_map[label] = best_target if best_ratio >= 0.82 and freq[best_target] >= freq[label] else label

    for s1 in stage1.values():
        ma = s1.get("major_area", "")
        if ma in label_map:
            s1["major_area"] = label_map[ma]


def _build_analysis_result(
    tables: Dict[str, SimpleTable],
    stage1: Dict[str, Dict[str, Any]],
    file_sources: Dict[str, str] | None = None,
) -> tuple[list, list]:
    """Returns (groups, not_openable). Groups are sorted by major_area."""
    group_map: Dict[str, list] = defaultdict(list)
    not_openable: list = []

    for key, table in tables.items():
        s1 = stage1.get(key, {})
        if not s1:
            continue
        table_display = table.table_kr or table.table_en or key
        row: Dict[str, Any] = {
            "table": table_display,
            "key": key,
            "bucket": s1.get("bucket", ""),
            "open_columns": s1.get("open_columns", []),
            "closed_columns": s1.get("closed_columns", []),
            "open_count": s1.get("open_count", 0),
            "total_count": s1.get("total_count", len(table.columns)),
            "major_area": s1.get("major_area", ""),
            "sub_area": s1.get("sub_area", ""),
            "dataset_name": s1.get("dataset_name", ""),
        }
        if file_sources and key in file_sources:
            row["source_file"] = file_sources[key]

        bucket = s1.get("bucket", "")
        if bucket in ("전체개방", "부분개방"):
            major = s1.get("major_area") or "미분류"
            group_map[major].append(row)
        else:
            not_openable.append(row)

    groups = [
        {"major_area": major, "tables": rows}
        for major, rows in sorted(group_map.items())
    ]
    return groups, not_openable


def _serialize_tables(tables: Dict[str, SimpleTable]) -> Dict[str, Dict[str, Any]]:
    return {
        key: {
            "key": t.key,
            "table_kr": t.table_kr,
            "table_en": t.table_en,
            "columns": list(t.columns),
            "file_sources": list(getattr(t, "file_sources", []) or []),
        }
        for key, t in tables.items()
    }


_REASON_LABELS = {
    1: "법률상 비공개",
    2: "국가안보·외교",
    3: "국민 생명·재산",
    4: "재판·수사",
    5: "감사·검사·인사",
    6: "개인정보",
    7: "영업비밀",
    8: "부동산투기",
    9: "시스템 테이블",
}


def _split_col_name(col_display: str) -> tuple[str, str]:
    """`태그(TAG)` → (`태그`, `TAG`); `태그` → (`태그`, ``); `TAG` → (``, `TAG`)."""
    if not col_display:
        return "", ""
    s = str(col_display).strip()
    if "(" in s and s.endswith(")"):
        idx = s.rfind("(")
        kr = s[:idx].strip()
        en = s[idx + 1 : -1].strip()
        return kr, en
    has_hangul = any("가" <= ch <= "힣" for ch in s)
    return (s, "") if has_hangul else ("", s)


def _format_reason(closed: Dict[str, Any]) -> str:
    """[6] 개인정보  /  [6, 7] 개인정보, 영업비밀  형태로 포맷."""
    codes = closed.get("reason_codes") or []
    text = (closed.get("reason") or "").strip()
    valid_codes = []
    for c in codes:
        try:
            n = int(c)
            if 1 <= n <= 9 and n not in valid_codes:
                valid_codes.append(n)
        except (TypeError, ValueError):
            continue
    if valid_codes:
        labels = ", ".join(f"{n}.{_REASON_LABELS.get(n, '?')}" for n in valid_codes)
        return f"{labels} ({text})" if text and text not in labels else labels
    return text


def _expand_columns(s1: Dict[str, Any]) -> List[tuple[str, str, str]]:
    """Return [(컬럼한글, 컬럼영문, 사유)] for both open and closed columns."""
    rows: List[tuple[str, str, str]] = []
    for name in s1.get("open_columns", []):
        kr, en = _split_col_name(name)
        rows.append((kr, en, ""))
    for c in s1.get("closed_columns", []):
        kr, en = _split_col_name(c.get("name", ""))
        rows.append((kr, en, _format_reason(c)))
    return rows


def _build_excel_workbook(
    tables_serialized: Dict[str, Dict[str, Any]],
    stage1: Dict[str, Dict[str, Any]],
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    HEADER_FILL = PatternFill("solid", fgColor="F0F1F4")
    HEADER_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _style_header(ws, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    def _hierarchical_merge(ws, last_row: int, max_col: int) -> None:
        """좌측 max_col개 컬럼을 계층적으로 병합.

        col=1은 col1 값이 같은 연속 행 → 대분류가 같은 모든 테이블이 묶임.
        col=2는 (col1, col2) 같은 연속 행 → 같은 대분류 내 같은 소분류 묶음.
        col=N은 (col1..colN) 모두 같은 연속 행 → 자연스럽게 테이블 단위로 좁아짐.
        """
        if last_row < 3:
            return

        # 병합 시 좌상단 외 셀이 클리어되므로 사전에 값 스냅샷
        snapshot: Dict[int, tuple] = {}
        for r in range(2, last_row + 1):
            snapshot[r] = tuple(
                "" if ws.cell(row=r, column=c).value is None else str(ws.cell(row=r, column=c).value)
                for c in range(1, max_col + 1)
            )

        for col in range(1, max_col + 1):
            run_start = 2
            for r in range(3, last_row + 2):
                if r > last_row:
                    same = False
                else:
                    same = all(
                        snapshot[r][c - 1] == snapshot[run_start][c - 1]
                        for c in range(1, col + 1)
                    )
                if not same:
                    if r - 1 > run_start:
                        ws.merge_cells(
                            start_row=run_start, start_column=col,
                            end_row=r - 1, end_column=col,
                        )
                        ws.cell(row=run_start, column=col).alignment = CENTER
                    run_start = r

    wb = Workbook()
    wb.remove(wb.active)

    # ── 개방가능 데이터셋 (전체개방 + 부분개방) ──
    open_headers = [
        "주제(대분류)", "주제(소분류)", "테이블명(한글)", "테이블명(영문)",
        "판정", "데이터셋명", "개방가능컬럼수", "전체컬럼수",
        "컬럼명(한글)", "컬럼명(영문)", "사유",
    ]
    open_widths = [12, 12, 18, 18, 9, 22, 9, 9, 16, 16, 22]
    ws_open = wb.create_sheet("개방가능_데이터셋")
    ws_open.append(open_headers)
    _style_header(ws_open, len(open_headers))
    for i, w in enumerate(open_widths, start=1):
        ws_open.column_dimensions[ws_open.cell(row=1, column=i).column_letter].width = w

    next_row = 2
    # 시각적으로 그룹이 묶이도록 (대분류, 소분류, 테이블명) 순서로 정렬해서 추가
    open_entries = sorted(
        (
            (key, tables_serialized[key], stage1[key])
            for key in tables_serialized
            if stage1.get(key, {}).get("bucket") in ("전체개방", "부분개방")
        ),
        key=lambda x: (
            x[2].get("major_area", ""),
            x[2].get("sub_area", ""),
            x[1].get("table_kr", "") or x[1].get("table_en", ""),
        ),
    )
    for key, table_dict, s1 in open_entries:
        col_rows = _expand_columns(s1) or [("", "", "")]
        for kr, en, reason in col_rows:
            ws_open.append([
                s1.get("major_area", ""),
                s1.get("sub_area", ""),
                table_dict.get("table_kr", ""),
                table_dict.get("table_en", ""),
                s1.get("bucket", ""),
                s1.get("dataset_name", ""),
                s1.get("open_count", 0),
                s1.get("total_count", 0),
                kr,
                en,
                reason,
            ])
            for c in range(1, 12):
                ws_open.cell(row=next_row, column=c).alignment = LEFT_TOP
            next_row += 1
    _hierarchical_merge(ws_open, last_row=next_row - 1, max_col=8)

    # ── 개방불가 목록 ──
    no_headers = [
        "테이블명(한글)", "테이블명(영문)",
        "컬럼명(한글)", "컬럼명(영문)", "사유",
    ]
    no_widths = [20, 20, 16, 16, 24]
    ws_no = wb.create_sheet("개방불가_목록")
    ws_no.append(no_headers)
    _style_header(ws_no, len(no_headers))
    for i, w in enumerate(no_widths, start=1):
        ws_no.column_dimensions[ws_no.cell(row=1, column=i).column_letter].width = w

    next_row = 2
    no_entries = sorted(
        (
            (key, tables_serialized[key], stage1[key])
            for key in tables_serialized
            if stage1.get(key, {}).get("bucket") == "불가능"
        ),
        key=lambda x: x[1].get("table_kr", "") or x[1].get("table_en", ""),
    )
    for key, table_dict, s1 in no_entries:
        col_rows = _expand_columns(s1) or [("", "", "")]
        for kr, en, reason in col_rows:
            ws_no.append([
                table_dict.get("table_kr", ""),
                table_dict.get("table_en", ""),
                kr,
                en,
                reason,
            ])
            for c in range(1, 6):
                ws_no.cell(row=next_row, column=c).alignment = LEFT_TOP
            next_row += 1
    _hierarchical_merge(ws_no, last_row=next_row - 1, max_col=2)

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name


class OpenDataAnalyzerService:

    def _get_api_key(self) -> str | None:
        return os.environ.get("GEMINI_API_KEY") or None

    def _get_base_url(self) -> str:
        return os.environ.get("GEMINI_BASE_URL") or GEMINI_BASE_URL

    def _get_model(self) -> str:
        return os.environ.get("GEMINI_MODEL") or GEMINI_MODEL

    async def start_stage1_background(
        self,
        files_data: List[tuple[bytes, str]],
        session_id: str | None = None,
        mock: bool = False,
        db: AsyncSession | None = None,
    ) -> dict:
        """Parse files, create execution record, launch background analysis.
        Returns immediately with {execution_id, session_id, status: 'running'}.
        """
        api_key = self._get_api_key()
        if not api_key and not mock:
            raise ValueError("Missing GEMINI_API_KEY.")

        new_file_names = [name for _, name in files_data]
        is_new_session = not (session_id and session_id in SESSIONS)

        if not is_new_session:
            state = SESSIONS[session_id]
            tables = state["tables"]
            file_sources = state.get("file_sources", {})
        else:
            if not files_data:
                raise ValueError("At least one file is required.")

            all_tables: Dict[str, SimpleTable] = {}
            file_sources: Dict[str, str] = {}

            # Phase 1: 모든 파일에서 테이블 수집
            raw_entries: List[tuple[str, SimpleTable, str]] = []
            for file_bytes, filename in files_data:
                suffix = "." + filename.rsplit(".", 1)[-1] if "." in filename else ".xlsx"
                tmp_path = _save_upload_bytes(file_bytes, suffix)
                try:
                    file_tables = await asyncio.to_thread(
                        extract_tables_from_excel, tmp_path, None
                    )
                    for _orig_key, table in file_tables.items():
                        norm_key = normalize_table_key(table.table_kr, table.table_en)
                        raw_entries.append((norm_key, table, filename))
                finally:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass

            # Phase 2: 중복 집계 후 (n/m) 표기, 병합 없이 각자 독립 테이블로 등록
            key_totals: Dict[str, int] = {}
            for norm_key, _, _ in raw_entries:
                key_totals[norm_key] = key_totals.get(norm_key, 0) + 1
            key_instance: Dict[str, int] = {}
            for norm_key, table, filename in raw_entries:
                key_instance[norm_key] = key_instance.get(norm_key, 0) + 1
                n = key_instance[norm_key]
                total = key_totals[norm_key]
                if total > 1:
                    label = f" (중복 {n}/{total})"
                    if table.table_kr:
                        table.table_kr = table.table_kr + label
                    if table.table_en:
                        table.table_en = table.table_en + label
                    unique_key = f"{norm_key}::{n}"
                else:
                    unique_key = norm_key
                table.file_sources = [filename]
                all_tables[unique_key] = table
                file_sources[unique_key] = filename

            tables = all_tables
            session_id = uuid.uuid4().hex
            SESSIONS[session_id] = {
                "tables": tables,
                "file_sources": file_sources,
                "stage1": {},
            }
            state = SESSIONS[session_id]

        if db is not None:
            summary = (
                ", ".join(new_file_names)
                if new_file_names
                else f"세션 재실행 ({session_id[:8]})"
            )
            execution = await ExecutionRecorder.create(
                db,
                project_slug=PROJECT_SLUG,
                process_type="inprocess",
                input_metadata={
                    "file_names": new_file_names,
                    "mock": mock,
                    "session_id": session_id,
                    "is_new_session": is_new_session,
                },
                input_summary=summary,
            )
            exec_id = execution.execution_id
        else:
            exec_id = uuid.uuid4().hex

        _PROGRESS[exec_id] = {"done": 0, "total": len(tables), "status": "running"}

        asyncio.create_task(
            self._run_stage1_bg(
                exec_id=exec_id,
                session_id=session_id,
                tables=tables,
                file_sources=file_sources,
                state=state,
                mock=mock,
                api_key=api_key,
            )
        )

        return {"execution_id": exec_id, "session_id": session_id, "status": "running"}

    async def _run_stage1_bg(
        self,
        exec_id: str,
        session_id: str,
        tables: Dict[str, SimpleTable],
        file_sources: Dict[str, str],
        state: Dict[str, Any],
        mock: bool,
        api_key: str | None,
    ) -> None:
        """Background coroutine: run LLM analysis and persist result to DB."""
        def _progress_cb(done: int, total: int, failed: int = 0) -> None:
            _PROGRESS[exec_id] = {
                "done": done,
                "total": total,
                "failed": failed,
                "status": "running",
            }

        async with async_session_factory() as db:
            try:
                stage1, failed_calls = await asyncio.to_thread(
                    run_stage1,
                    tables,
                    REASONS,
                    api_key,
                    self._get_base_url(),
                    self._get_model(),
                    FIXED_SLEEP_SECONDS,
                    mock,
                    _progress_cb,
                )

                _consolidate_major_areas(stage1)
                state["stage1"].update(stage1)

                groups, not_openable = _build_analysis_result(tables, stage1, file_sources)

                full_open_count = sum(1 for r in stage1.values() if r.get("bucket") == "전체개방")
                partial_count = sum(1 for r in stage1.values() if r.get("bucket") == "부분개방")
                not_open_count = sum(1 for r in stage1.values() if r.get("bucket") == "불가능")

                response = {
                    "session_id": session_id,
                    "execution_id": exec_id,
                    "total": len(tables),
                    "full_open_count": full_open_count,
                    "partial_count": partial_count,
                    "not_openable_count": not_open_count,
                    "file_count": len(file_sources) if file_sources else 1,
                    "groups": groups,
                    "not_openable": not_openable,
                    "failed_count": len(failed_calls),
                    "failed": failed_calls,
                }
                state["result"] = response

                snapshot = {
                    "response": response,
                    "raw": {
                        "tables": _serialize_tables(tables),
                        "stage1": state["stage1"],
                        "file_sources": file_sources,
                        "failed": failed_calls,
                    },
                }
                await ExecutionRecorder.mark_succeeded(db, execution_id=exec_id, result_data=snapshot)

                total = _PROGRESS.get(exec_id, {}).get("total", len(tables))
                _PROGRESS[exec_id] = {
                    "done": total,
                    "total": total,
                    "failed": len(failed_calls),
                    "status": "succeeded",
                }

            except Exception as exc:
                logger.exception("Background stage1 failed exec_id=%s", exec_id)
                try:
                    await ExecutionRecorder.mark_failed(db, execution_id=exec_id, error_message=str(exc))
                except Exception:  # noqa: BLE001
                    pass
                _PROGRESS[exec_id] = {"done": 0, "total": 0, "status": "failed", "error": str(exc)}

    async def get_stage1_progress(self, execution_id: str, db: AsyncSession | None = None) -> dict:
        info = _PROGRESS.get(execution_id)
        if info:
            done = info.get("done", 0)
            total = info.get("total", 0)
            percent = round(done / total * 100) if total > 0 else 0
            return {
                "done": done, "total": total, "percent": percent,
                "failed": info.get("failed", 0),
                "status": info.get("status", "running"), "error": info.get("error"),
            }
        if db is not None:
            record = await ExecutionRecorder.get(db, execution_id)
            if record:
                is_done = record.status in ("succeeded", "failed")
                failed_count = 0
                if record.result_data:
                    failed_count = len(record.result_data.get("raw", {}).get("failed", []))
                return {
                    "done": 0, "total": 0, "percent": 100 if is_done else 0,
                    "failed": failed_count,
                    "status": record.status, "error": record.error_message,
                }
        return {"done": 0, "total": 0, "percent": 0, "failed": 0, "status": "unknown", "error": None}

    async def export_excel(self, session_id: str) -> str:
        if session_id not in SESSIONS:
            raise ValueError("Invalid session_id.")
        state = SESSIONS[session_id]
        tables = state["tables"]
        stage1 = state.get("stage1", {})
        return await asyncio.to_thread(
            _build_excel_workbook, _serialize_tables(tables), stage1
        )

    async def export_excel_from_execution(
        self,
        db: AsyncSession,
        execution_id: str,
    ) -> str:
        record = await ExecutionRecorder.get(db, execution_id)
        if record is None or record.project_slug != PROJECT_SLUG:
            raise ValueError("Invalid execution_id.")
        if record.status != "succeeded" or not record.result_data:
            raise ValueError("Execution has no result to export.")

        raw = record.result_data.get("raw", {})
        tables_serialized = raw.get("tables", {})
        stage1 = raw.get("stage1", {})

        return await asyncio.to_thread(
            _build_excel_workbook, tables_serialized, stage1
        )

    async def list_executions(
        self,
        db: AsyncSession,
        page: int = 1,
        page_size: int = 20,
    ) -> dict:
        result = await ExecutionRecorder.list(
            db,
            project_slug=PROJECT_SLUG,
            page=page,
            page_size=page_size,
        )
        return {
            "items": [serialize_execution(r) for r in result["items"]],
            "total": result["total"],
            "page": result["page"],
            "page_size": result["page_size"],
            "total_pages": result["total_pages"],
        }

    async def get_execution(
        self,
        db: AsyncSession,
        execution_id: str,
    ) -> dict | None:
        record = await ExecutionRecorder.get(db, execution_id)
        if record is None or record.project_slug != PROJECT_SLUG:
            return None
        return serialize_execution(record, include_result=True)

    async def delete_execution(
        self,
        db: AsyncSession,
        execution_id: str,
    ) -> bool:
        record = await ExecutionRecorder.get(db, execution_id)
        if record is None or record.project_slug != PROJECT_SLUG:
            return False
        return await ExecutionRecorder.delete(db, execution_id)

    async def get_stats(self) -> AnalyzerStats:
        return AnalyzerStats(
            active_sessions=len(SESSIONS),
            supported_formats=[".xlsx"],
            mock_available=True,
            stages=["1단계: 컬럼 단위 개방 가능 여부 판단 + 주제분류 + 데이터셋명"],
        )
