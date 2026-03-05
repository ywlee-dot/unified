"""
엑셀 파일 로드 유틸리티

각 단계별 엑셀 파일에서 워크시트를 찾고 데이터를 로드하기 위한
공통 함수들을 제공합니다.
"""

from typing import Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook


def find_stage_sheet(excel_path: str, stage_pattern: str) -> Optional[Any]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if stage_pattern in sheet_name:
            return wb[sheet_name]
    return None


def load_stage_sheet(excel_path: str, stage_pattern: str) -> Tuple[List[Tuple[Any, ...]], bool]:
    try:
        sheet = find_stage_sheet(excel_path, stage_pattern)
        if not sheet:
            return [], False
        rows = list(sheet.iter_rows(values_only=True))
        return rows, True
    except Exception:
        return [], False


def parse_stage1_row(row: Tuple[Any, ...]) -> Optional[Dict[str, Any]]:
    if not row or len(row) < 6:
        return None

    table_kr = str(row[0]).strip() if row[0] else ""
    table_en = str(row[1]).strip() if len(row) > 1 and row[1] else ""
    openable = str(row[3]).strip() if len(row) > 3 and row[3] else ""
    reason_nums_str = str(row[4]).strip() if len(row) > 4 and row[4] else ""
    reason_text = str(row[5]).strip() if len(row) > 5 and row[5] else ""

    if not table_kr and not table_en:
        return None

    reason_numbers = []
    if reason_nums_str:
        for num_str in reason_nums_str.split(","):
            try:
                reason_numbers.append(int(num_str.strip()))
            except ValueError:
                pass

    return {
        "table_kr": table_kr,
        "table_en": table_en,
        "openable": openable,
        "reason_numbers": reason_numbers,
        "reason_text": reason_text,
    }


def parse_stage2_row(row: Tuple[Any, ...]) -> Optional[Dict[str, Any]]:
    if not row or len(row) < 3:
        return None

    table_kr = str(row[0]).strip() if row[0] else ""
    table_en = str(row[1]).strip() if len(row) > 1 and row[1] else ""
    subject = str(row[2]).strip() if len(row) > 2 and row[2] else ""

    if not table_kr and not table_en:
        return None

    return {
        "table_kr": table_kr,
        "table_en": table_en,
        "subject_area": subject,
    }


def parse_stage3_row(row: Tuple[Any, ...]) -> Optional[Dict[str, Any]]:
    if not row or len(row) < 4:
        return None

    table_kr = str(row[0]).strip() if row[0] else ""
    table_en = str(row[1]).strip() if len(row) > 1 and row[1] else ""
    core_cols_str = str(row[2]).strip() if len(row) > 2 and row[2] else ""
    dataset_desc = str(row[3]).strip() if len(row) > 3 and row[3] else ""

    if not table_kr and not table_en:
        return None

    core_columns = core_cols_str.split(",") if core_cols_str else []

    return {
        "table_kr": table_kr,
        "table_en": table_en,
        "core_columns": [c.strip() for c in core_columns],
        "dataset_description": dataset_desc,
    }


def parse_stage4_row(row: Tuple[Any, ...]) -> Optional[Dict[str, Any]]:
    if not row or len(row) < 4:
        return None

    table_kr = str(row[0]).strip() if row[0] else ""
    table_en = str(row[1]).strip() if len(row) > 1 and row[1] else ""
    join_table = str(row[2]).strip() if len(row) > 2 and row[2] else ""
    join_keys_str = str(row[3]).strip() if len(row) > 3 and row[3] else ""

    if not table_kr and not table_en:
        return None

    join_keys = [k.strip() for k in join_keys_str.split(",")] if join_keys_str else []

    return {
        "table_kr": table_kr,
        "table_en": table_en,
        "join_table": join_table,
        "join_keys": join_keys,
    }
