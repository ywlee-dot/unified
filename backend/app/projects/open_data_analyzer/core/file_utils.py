"""
파일 I/O 유틸리티

엑셀 파일 처리를 위한 공통 함수들을 제공합니다.
"""

import re
from typing import Any, List, Optional, Tuple

from openpyxl import load_workbook


def read_sheet_rows(
    path: str, sheet_name: Optional[str] = None
) -> Tuple[List[str], List[Tuple[Any, ...]]]:
    wb = load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def find_col(headers: List[str], candidates: List[str]) -> Optional[int]:
    for cand in candidates:
        for idx, header in enumerate(headers):
            if header == cand:
                return idx

    for cand in candidates:
        for idx, header in enumerate(headers):
            if cand in header:
                return idx

    return None


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()
