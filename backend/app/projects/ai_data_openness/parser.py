"""
Excel 파서 모듈. 기존 스킬의 parse_excel.py를 웹 백엔드용으로 리팩토링.
파일 경로 대신 UploadFile 객체를 받을 수 있도록 확장.
"""

import io
import openpyxl

HEADER_KEYWORDS = [
    "데이터", "목록", "이름", "명칭", "분야", "형태", "유형",
    "설명", "URL", "개방", "제공", "항목", "번호", "No",
    "카테고리", "주제", "기관", "포맷", "갱신", "주기",
    "테이블", "컬럼", "시스템", "비고", "정보", "데이터셋",
]

COLUMN_MAPPING = {
    "데이터명": ["데이터명", "데이터 명", "목록명", "데이터셋명", "데이터셋 명",
                "데이터이름", "이름", "명칭", "제목", "개방 데이터셋명"],
    "분야": ["분야", "카테고리", "주제", "주제분류", "분류", "대분류"],
    "설명": ["설명", "내용", "요약", "데이터설명", "데이터 설명", "상세설명"],
    "형태": ["형태", "유형", "포맷", "제공형태", "데이터형태", "파일형태", "제공포맷", "확장자"],
    "개방시점": ["개방시점", "개방시기", "개방일", "개방일자", "등록일", "등록일자", "공개일", "제공일"],
    "URL": ["URL", "url", "링크", "개방URL", "다운로드", "주소", "개방주소"],
    "기관명": ["기관명", "기관", "제공기관", "소관기관", "담당기관"],
    "갱신주기": ["갱신주기", "갱신 주기", "업데이트주기", "주기", "갱신"],
    "비고": ["비고", "참고", "상태", "메모", "기타"],
    "시스템명": ["정보시스템명", "시스템명", "시스템", "소스시스템"],
}


def normalize_column_name(name: str) -> str:
    name = name.strip()
    for standard, variants in COLUMN_MAPPING.items():
        for v in variants:
            if v == name or v in name:
                return standard
    return name


def choose_best_sheet(wb) -> str:
    skip_keywords = ["표지", "목차", "안내", "설명", "참고"]
    best_sheet, best_rows = None, 0
    for name in wb.sheetnames:
        if any(k in name for k in skip_keywords):
            continue
        ws = wb[name]
        if ws.max_row and ws.max_row > best_rows:
            best_rows = ws.max_row
            best_sheet = name
    return best_sheet or wb.sheetnames[0]


def detect_header_row(ws, max_scan=15):
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]
        non_empty = [v for v in row_values if v]
        if len(non_empty) < 2:
            continue
        match_count = sum(1 for v in non_empty for k in HEADER_KEYWORDS if k in v)
        if match_count >= 2:
            return row_idx, row_values
    return 1, [str(cell.value or "").strip() for cell in ws[1]]


def detect_mode(ws, header_row_idx: int) -> str:
    merged_count = len(ws.merged_cells.ranges)
    data_rows = ws.max_row - header_row_idx
    if data_rows <= 0:
        return "flat"
    if merged_count / max(data_rows, 1) > 0.3 or merged_count > 20:
        return "hierarchical"
    return "flat"


def parse_flat(ws) -> dict:
    header_row_idx, raw_headers = detect_header_row(ws)
    headers = [normalize_column_name(h) for h in raw_headers]
    valid_indices = [i for i, h in enumerate(headers) if h]

    data_list = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_cells = list(ws[row_idx])
        row_data = {}
        all_empty = True
        for i in valid_indices:
            if i < len(row_cells):
                val = row_cells[i].value
                if val is not None:
                    all_empty = False
                    row_data[headers[i]] = str(val).strip()
                else:
                    row_data[headers[i]] = ""
        if not all_empty:
            data_list.append(row_data)

    return {"mode": "flat", "columns": [headers[i] for i in valid_indices],
            "total_count": len(data_list), "data": data_list}


def find_id_and_name_cols(ws, header_row_idx):
    header_row = [str(cell.value or "").strip() for cell in ws[header_row_idx]]
    sub_header = [""] * len(header_row)
    if header_row_idx + 1 <= ws.max_row:
        sub_header = [str(cell.value or "").strip() for cell in ws[header_row_idx + 1]]
    merged = [h or s for h, s in zip(header_row, sub_header)]

    id_col, name_col = None, None
    for i, h in enumerate(merged):
        if id_col is None and any(k in h for k in ["번호", "No", "순번", "연번"]):
            id_col = i
        if name_col is None and any(k in h for k in ["데이터셋", "데이터명", "목록명"]):
            name_col = i

    if id_col is not None and name_col is None:
        name_col = id_col + 1

    if id_col is None:
        for row_idx in range(header_row_idx + 1, min(header_row_idx + 10, ws.max_row + 1)):
            cells = list(ws[row_idx])
            for i, cell in enumerate(cells):
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    return i, i + 1
    return id_col or 0, name_col or 1


def parse_hierarchical(ws) -> dict:
    header_row_idx, _ = detect_header_row(ws)
    id_col, name_col = find_id_and_name_cols(ws, header_row_idx)

    header1 = [str(cell.value or "").strip() for cell in ws[header_row_idx]]
    start_row = header_row_idx + 1
    if start_row <= ws.max_row:
        header2 = [str(cell.value or "").strip() for cell in ws[start_row]]
        has_data = any(isinstance(cell.value, (int, float)) for cell in ws[start_row] if cell.value)
        if not has_data and sum(1 for h in header2 if h) >= 2:
            start_row += 1
        else:
            header2 = [""] * len(header1)
    else:
        header2 = [""] * len(header1)

    col_kr_idx, col_en_idx = None, None
    system_idx, open_time_idx, note_idx = None, None, None
    for i, (h1, h2) in enumerate(zip(header1, header2)):
        combined = h1 + h2
        if "컬럼" in combined and ("한글" in combined or "명" in combined) and "영" not in combined:
            col_kr_idx = i
        elif "컬럼" in combined and "영" in combined:
            col_en_idx = i
        if any(k in combined for k in ["시스템", "정보시스템"]):
            system_idx = i
        elif any(k in combined for k in ["개방시기", "개방시점"]):
            open_time_idx = i
        elif any(k in combined for k in ["비고", "상태"]):
            note_idx = i

    datasets = []
    current = None
    for row_idx in range(start_row, ws.max_row + 1):
        cells = list(ws[row_idx])
        if id_col >= len(cells):
            continue
        id_val = cells[id_col].value
        name_val = cells[name_col].value if name_col < len(cells) else None

        if id_val is not None and str(id_val).strip():
            current = {
                "번호": str(id_val).strip(),
                "데이터명": str(name_val or "").strip(),
                "시스템명": str(cells[system_idx].value or "").strip() if system_idx and system_idx < len(cells) else "",
                "개방시점": str(cells[open_time_idx].value or "").strip().replace("\n", " ") if open_time_idx and open_time_idx < len(cells) else "",
                "비고": str(cells[note_idx].value or "").strip() if note_idx and note_idx < len(cells) else "",
                "컬럼목록": [],
            }
            datasets.append(current)

        if current and col_kr_idx and col_kr_idx < len(cells):
            col_kr_val = cells[col_kr_idx].value
            if col_kr_val:
                col_info = {"한글명": str(col_kr_val).strip()}
                if col_en_idx and col_en_idx < len(cells) and cells[col_en_idx].value:
                    col_info["영문명"] = str(cells[col_en_idx].value).strip()
                current["컬럼목록"].append(col_info)

    for d in datasets:
        d["컬럼수"] = len(d["컬럼목록"])
        col_names = [c["한글명"] for c in d["컬럼목록"]]
        d["컬럼명_요약"] = ", ".join(col_names[:8])
        if len(col_names) > 8:
            d["컬럼명_요약"] += f" 외 {len(col_names)-8}개"

    return {"mode": "hierarchical", "total_count": len(datasets), "data": datasets}


def parse_excel(file_content: bytes, filename: str = "") -> dict:
    """바이트 데이터로부터 Excel을 파싱. 웹 업로드용."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    sheet_name = choose_best_sheet(wb)
    ws = wb[sheet_name]

    header_row_idx, _ = detect_header_row(ws)
    mode = detect_mode(ws, header_row_idx)

    if mode == "hierarchical":
        result = parse_hierarchical(ws)
    else:
        result = parse_flat(ws)

    result["filename"] = filename
    result["sheet"] = sheet_name
    wb.close()
    return result
