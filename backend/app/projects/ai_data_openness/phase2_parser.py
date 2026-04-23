"""
Phase 2 Excel 파서 모듈. '발굴 목록' 형식의 Excel 파일을 파싱.
Sheet: "AI 친화·고가치 데이터 발굴 목록"
Header row: 2 (번호 | 발굴 루트 | 담당부서 | 후보군명 | 구성안 | 예상활용사례 | 비고 | 개방가능여부)
"""

import io
import openpyxl

PHASE2_SKIP_KEYWORDS = ["표지", "목차", "안내", "설명", "참고"]

PHASE2_TARGET_SHEET_KEYWORDS = ["발굴", "AI 친화", "고가치"]

PHASE2_COLUMN_MAPPING = {
    "번호": ["번호", "No", "no", "순번", "연번"],
    "발굴루트": ["발굴 루트", "발굴루트", "루트", "발굴경로"],
    "담당부서": ["담당부서", "담당 부서", "부서", "소관부서", "담당기관"],
    "후보군명": [
        "AI 친화·고가치 데이터 후보군명(가칭)",
        "AI 친화·고가치 데이터 후보군명",
        "후보군명",
        "후보군",
        "데이터명",
        "데이터 명",
    ],
    "구성안": [
        "AI 친화·고가치 데이터 구성(안)",
        "AI 친화·고가치 데이터 구성안",
        "구성(안)",
        "구성안",
        "데이터 구성",
    ],
    "예상활용사례": ["예상 활용 사례", "예상활용사례", "활용사례", "활용 사례"],
    "비고": ["비고", "참고", "메모", "기타"],
    "개방가능여부": ["개방 가능 여부", "개방가능여부", "개방여부", "개방 여부"],
}

PHASE2_HEADER_KEYWORDS = [
    "번호", "발굴", "담당", "부서", "후보군", "구성", "활용", "개방", "비고",
]


def normalize_phase2_column(name: str) -> str:
    name = name.strip()
    for standard, variants in PHASE2_COLUMN_MAPPING.items():
        for v in variants:
            if v == name or v in name:
                return standard
    return name


def choose_phase2_sheet(wb) -> str:
    """발굴 목록 시트를 우선 선택, 없으면 표지/목차/안내를 제외한 가장 큰 시트."""
    # Try to find a sheet explicitly matching target keywords
    for name in wb.sheetnames:
        if any(k in name for k in PHASE2_TARGET_SHEET_KEYWORDS):
            return name

    # Fall back: skip known non-data sheets, pick the one with most rows
    best_sheet, best_rows = None, 0
    for name in wb.sheetnames:
        if any(k in name for k in PHASE2_SKIP_KEYWORDS):
            continue
        ws = wb[name]
        if ws.max_row and ws.max_row > best_rows:
            best_rows = ws.max_row
            best_sheet = name

    return best_sheet or wb.sheetnames[0]


def detect_phase2_header_row(ws, max_scan: int = 10):
    """헤더 행 자동 감지. 2개 이상의 Phase 2 키워드가 매칭되는 행을 반환."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]
        non_empty = [v for v in row_values if v]
        if len(non_empty) < 2:
            continue
        match_count = sum(
            1 for v in non_empty for k in PHASE2_HEADER_KEYWORDS if k in v
        )
        if match_count >= 2:
            return row_idx, row_values
    # Default to row 1
    return 1, [str(cell.value or "").strip() for cell in ws[1]]


def clean_cell(value) -> str:
    """셀 값을 문자열로 변환하고 공백을 정리."""
    if value is None:
        return ""
    return str(value).strip()


def parse_phase2_excel(file_content: bytes, filename: str) -> dict:
    """
    Phase 2 '발굴 목록' 형식의 Excel을 파싱.

    Returns:
        {
            "filename": str,
            "sheet": str,
            "total_count": int,
            "data": [
                {
                    "번호": str,
                    "발굴루트": str,
                    "담당부서": str,
                    "후보군명": str,
                    "구성안": str,
                    "예상활용사례": str,
                    "비고": str,
                    "개방가능여부": str,
                },
                ...
            ]
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    sheet_name = choose_phase2_sheet(wb)
    ws = wb[sheet_name]

    header_row_idx, raw_headers = detect_phase2_header_row(ws)
    headers = [normalize_phase2_column(h) for h in raw_headers]

    # Build index map: normalized column name → column index (first occurrence wins)
    col_index: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h and h not in col_index:
            col_index[h] = i

    expected_columns = ["번호", "발굴루트", "담당부서", "후보군명", "구성안", "예상활용사례", "비고", "개방가능여부"]

    data_list = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_cells = list(ws[row_idx])

        # Check if row is entirely empty
        all_empty = all(
            (row_cells[i].value is None or str(row_cells[i].value).strip() == "")
            for i in range(len(row_cells))
        )
        if all_empty:
            continue

        row_data = {}
        for col_name in expected_columns:
            idx = col_index.get(col_name)
            if idx is not None and idx < len(row_cells):
                row_data[col_name] = clean_cell(row_cells[idx].value)
            else:
                row_data[col_name] = ""

        data_list.append(row_data)

    wb.close()

    return {
        "filename": filename,
        "sheet": sheet_name,
        "total_count": len(data_list),
        "data": data_list,
    }
