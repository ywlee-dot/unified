"""Business logic service for Dataset Summary project."""

from __future__ import annotations

import asyncio
import logging
import os
import tempfile
import uuid
from typing import Any, Dict, List

from sqlalchemy.ext.asyncio import AsyncSession

from app.database import async_session_factory
from app.projects.dataset_summary.core import (
    DEFAULT_BASE_URL,
    DEFAULT_MODEL,
    generate_mock_summaries,
    generate_summaries,
    read_csv_rows_from_bytes,
    read_rows_from_bytes,
)
from app.projects.dataset_summary.schemas import DatasetSummaryStats
from app.shared.services.execution_recorder import (
    ExecutionRecorder,
    serialize_execution,
)

logger = logging.getLogger(__name__)

PROJECT_SLUG = "dataset_summary"

# In-memory progress tracker: execution_id → {done, total, status, error?}
_PROGRESS: Dict[str, Dict[str, Any]] = {}


def _build_summary_workbook(results: List[Dict[str, Any]]) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "데이터셋 설명·키워드"

    headers = ["#", "대분류", "소분류", "데이터셋명", "테이블명", "키워드", "설명", "컬럼 목록"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0064FF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in results:
        common = item.get("common", {}) or {}
        keywords = item.get("keywords", []) or []
        columns = item.get("columns", []) or []
        ws.append([
            item.get("row_index", ""),
            common.get("대분류", ""),
            common.get("소분류", ""),
            common.get("데이터셋명", ""),
            common.get("테이블명", ""),
            ", ".join(str(k) for k in keywords),
            item.get("description", ""),
            ", ".join(str(c) for c in columns),
        ])

    widths = [5, 14, 14, 26, 22, 36, 60, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    handle.close()
    wb.save(handle.name)
    return handle.name


class DatasetSummaryService:
    """Service for generating dataset summaries via LLM."""

    async def start_summarize_background(
        self,
        file_bytes: bytes,
        filename: str,
        sheet: str | None = None,
        org_name: str | None = None,
        include_rows: bool = False,
        mock: bool = False,
        include_prompt: bool = False,
        db: AsyncSession | None = None,
    ) -> dict:
        """Parse file, create execution record, launch background generation.
        Returns immediately with {execution_id, status: 'running', total}.
        """
        use_mock = mock or os.environ.get("MOCK_MODE") == "1"
        api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("OPENAI_API_KEY")
        if not api_key and not use_mock:
            raise ValueError("Missing GEMINI_API_KEY/OPENAI_API_KEY.")

        if filename.endswith(".xlsx"):
            rows = await asyncio.to_thread(read_rows_from_bytes, file_bytes, sheet)
        elif filename.endswith(".csv"):
            rows = await asyncio.to_thread(read_csv_rows_from_bytes, file_bytes)
        else:
            raise ValueError("Only .xlsx or .csv files are supported.")

        if not rows:
            raise ValueError("No rows found.")

        from app.projects.dataset_summary.core import group_rows
        groups = group_rows(rows)
        total = len(groups)

        if db is not None:
            execution = await ExecutionRecorder.create(
                db,
                project_slug=PROJECT_SLUG,
                process_type="inprocess",
                input_metadata={
                    "filename": filename,
                    "sheet": sheet,
                    "org_name": org_name,
                    "mock": use_mock,
                    "total": total,
                },
                input_summary=f"{org_name or '-'} · {filename}",
            )
            exec_id = execution.execution_id
        else:
            exec_id = uuid.uuid4().hex

        _PROGRESS[exec_id] = {"done": 0, "total": total, "status": "running"}

        asyncio.create_task(
            self._run_summarize_bg(
                exec_id=exec_id,
                rows=rows,
                org_name=org_name,
                include_rows=include_rows,
                include_prompt=include_prompt,
                use_mock=use_mock,
                api_key=api_key,
            )
        )

        return {"execution_id": exec_id, "status": "running", "total": total}

    async def _run_summarize_bg(
        self,
        exec_id: str,
        rows: List[Dict[str, Any]],
        org_name: str | None,
        include_rows: bool,
        include_prompt: bool,
        use_mock: bool,
        api_key: str | None,
    ) -> None:
        """Background coroutine: run LLM generation and persist result to DB."""
        def _progress_cb(done: int, total: int, failed: int = 0) -> None:
            _PROGRESS[exec_id] = {
                "done": done,
                "total": total,
                "failed": failed,
                "status": "running",
            }

        async with async_session_factory() as db:
            try:
                if use_mock:
                    results = await asyncio.to_thread(
                        generate_mock_summaries,
                        rows,
                        include_rows=include_rows,
                        include_prompt=include_prompt,
                        org_name=org_name,
                    )
                else:
                    results = await asyncio.to_thread(
                        generate_summaries,
                        rows,
                        DEFAULT_BASE_URL, api_key, DEFAULT_MODEL,
                        include_rows=include_rows,
                        include_prompt=include_prompt,
                        org_name=org_name,
                        progress_callback=_progress_cb,
                    )

                response: Dict[str, Any] = {"results": results, "execution_id": exec_id}

                snapshot = {
                    "response": response,
                    "options": {
                        "include_rows": include_rows,
                        "include_prompt": include_prompt,
                    },
                }
                await ExecutionRecorder.mark_succeeded(
                    db, execution_id=exec_id, result_data=snapshot
                )

                total = _PROGRESS.get(exec_id, {}).get("total", len(results))
                _PROGRESS[exec_id] = {
                    "done": total,
                    "total": total,
                    "status": "succeeded",
                }
            except Exception as exc:
                logger.exception("Background summarize failed exec_id=%s", exec_id)
                try:
                    await ExecutionRecorder.mark_failed(
                        db, execution_id=exec_id, error_message=str(exc)
                    )
                except Exception:  # noqa: BLE001
                    pass
                _PROGRESS[exec_id] = {
                    "done": 0,
                    "total": 0,
                    "status": "failed",
                    "error": str(exc),
                }

    async def export_excel_from_execution(
        self,
        db: AsyncSession,
        execution_id: str,
    ) -> str:
        record = await ExecutionRecorder.get(db, execution_id)
        if record is None or record.project_slug != PROJECT_SLUG:
            raise ValueError("Invalid execution_id.")
        if record.status != "succeeded" or not record.result_data:
            raise ValueError("Execution has no result to export.")

        results = (
            record.result_data.get("response", {}).get("results", [])
            if isinstance(record.result_data, dict)
            else []
        )
        if not results:
            raise ValueError("결과가 비어 있습니다.")

        return await asyncio.to_thread(_build_summary_workbook, results)

    async def get_summarize_progress(
        self,
        execution_id: str,
        db: AsyncSession | None = None,
    ) -> dict:
        info = _PROGRESS.get(execution_id)
        if info:
            done = info.get("done", 0)
            total = info.get("total", 0)
            percent = round(done / total * 100) if total > 0 else 0
            return {
                "done": done,
                "total": total,
                "percent": percent,
                "status": info.get("status", "running"),
                "error": info.get("error"),
            }
        if db is not None:
            record = await ExecutionRecorder.get(db, execution_id)
            if record:
                is_done = record.status in ("succeeded", "failed")
                return {
                    "done": 0,
                    "total": 0,
                    "percent": 100 if is_done else 0,
                    "status": record.status,
                    "error": record.error_message,
                }
        return {
            "done": 0,
            "total": 0,
            "percent": 0,
            "status": "unknown",
            "error": None,
        }

    # ── Execution history ─────────────────────────────────────────────────

    async def list_executions(
        self,
        db: AsyncSession,
        page: int = 1,
        page_size: int = 20,
    ) -> dict:
        result = await ExecutionRecorder.list(
            db,
            project_slug=PROJECT_SLUG,
            page=page,
            page_size=page_size,
        )
        return {
            "items": [serialize_execution(r) for r in result["items"]],
            "total": result["total"],
            "page": result["page"],
            "page_size": result["page_size"],
            "total_pages": result["total_pages"],
        }

    async def get_execution(
        self,
        db: AsyncSession,
        execution_id: str,
    ) -> dict | None:
        record = await ExecutionRecorder.get(db, execution_id)
        if record is None or record.project_slug != PROJECT_SLUG:
            return None
        return serialize_execution(record, include_result=True)

    async def delete_execution(
        self,
        db: AsyncSession,
        execution_id: str,
    ) -> bool:
        record = await ExecutionRecorder.get(db, execution_id)
        if record is None or record.project_slug != PROJECT_SLUG:
            return False
        return await ExecutionRecorder.delete(db, execution_id)

    async def get_stats(self) -> DatasetSummaryStats:
        return DatasetSummaryStats(
            total_generated=0,
            mock_available=True,
            supported_formats=[".xlsx", ".csv"],
        )
