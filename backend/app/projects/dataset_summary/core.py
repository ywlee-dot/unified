import json
import os
import re
import time
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List

import csv
import requests
from openpyxl import load_workbook

DEFAULT_BASE_URL = (
    os.environ.get("OPENAI_BASE_URL")
    or os.environ.get("GEMINI_BASE_URL")
    or "https://generativelanguage.googleapis.com/v1beta"
)
DEFAULT_MODEL = (
    os.environ.get("OPENAI_MODEL")
    or os.environ.get("GEMINI_MODEL")
    or "gemini-2.5-flash"
)
DEFAULT_PROMPT_PATH = os.environ.get("PROMPT_PATH", str(Path(__file__).parent / "prompt.txt"))

REQUIRED_COLUMNS = ["대분류", "소분류", "데이터셋명", "테이블명", "컬럼명"]
HIERARCHY = ["대분류", "소분류", "데이터셋명", "테이블명"]
COMMON_KEYS = ("대분류", "소분류", "데이터셋명", "테이블명")

HEADER_ALIASES: Dict[str, str] = {
    "데이터셋이름": "데이터셋명",
    "테이블이름": "테이블명",
    "컬럼이름": "컬럼명",
    "ds명": "데이터셋명",
    "주제(대분류)": "대분류",
    "주제(소분류)": "소분류",
}

_PAREN_RE = re.compile(r"[\(（][^)）]*[\)）]")
_PROMPT_CACHE = None


def _normalize_header(s: str) -> str:
    return "".join((s or "").split()).lower()


def _strip_paren(s: str) -> str:
    return _PAREN_RE.sub("", s).strip()


def _canonical_header(name: str) -> str | None:
    if not name:
        return None
    norm = _normalize_header(name)
    for canonical in REQUIRED_COLUMNS:
        if _normalize_header(canonical) == norm:
            return canonical
    if norm in HEADER_ALIASES:
        return HEADER_ALIASES[norm]

    stripped = _normalize_header(_strip_paren(name))
    if stripped and stripped != norm:
        for canonical in REQUIRED_COLUMNS:
            if _normalize_header(canonical) == stripped:
                return canonical
        if stripped in HEADER_ALIASES:
            return HEADER_ALIASES[stripped]

    return None


def _find_header_row(rows: List[tuple], max_scan: int = 50) -> int | None:
    for i, row in enumerate(rows[:max_scan]):
        canonicals = set()
        for c in row:
            cell = str(c).strip() if c is not None else ""
            canonical = _canonical_header(cell)
            if canonical:
                canonicals.add(canonical)
        if all(col in canonicals for col in REQUIRED_COLUMNS):
            return i
    return None


def _rows_to_dicts(rows: List[tuple]) -> List[Dict[str, Any]]:
    if not rows:
        return []

    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError(
            f"필수 헤더 컬럼을 찾을 수 없습니다. 파일에 다음 컬럼이 있어야 합니다: {', '.join(REQUIRED_COLUMNS)}"
        )

    raw_header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    col_map: Dict[str, int] = {}
    for i, name in enumerate(raw_header):
        canonical = _canonical_header(name)
        if canonical and canonical not in col_map:
            col_map[canonical] = i

    last_seen: Dict[str, str] = {}
    results: List[Dict[str, Any]] = []

    for row in rows[header_idx + 1:]:
        values: Dict[str, str] = {}
        for col_name, idx in col_map.items():
            v = row[idx] if idx < len(row) else None
            values[col_name] = str(v).strip() if v is not None else ""

        for level, col in enumerate(HIERARCHY):
            if col not in col_map:
                continue
            v = values.get(col, "")
            if v:
                last_seen[col] = v
                for descendant in HIERARCHY[level + 1:]:
                    last_seen.pop(descendant, None)
            elif col in last_seen:
                values[col] = last_seen[col]

        row_dict: Dict[str, Any] = {}
        for col in REQUIRED_COLUMNS:
            if values.get(col):
                row_dict[col] = values[col]

        if row_dict:
            results.append(row_dict)

    return results


def read_rows_from_bytes(
    xlsx_bytes: bytes,
    sheet_name: str | None,
) -> List[Dict[str, Any]]:
    workbook = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return _rows_to_dicts(rows)


def read_csv_rows_from_bytes(csv_bytes: bytes) -> List[Dict[str, Any]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = csv_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Unable to decode CSV. Try UTF-8 or CP949.")

    reader = csv.reader(StringIO(text))
    rows = [tuple(r) for r in reader]
    return _rows_to_dicts(rows)


def group_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: List[Dict[str, Any]] = []
    current_key: str | None = None
    current_rows: List[Dict[str, Any]] = []

    for row in rows:
        key_value = row.get("데이터셋명", "").strip()
        if not key_value:
            if current_rows:
                current_rows.append(row)
            continue
        if current_key is None:
            current_key = key_value
            current_rows = [row]
        elif key_value != current_key:
            groups.append({"group_key": current_key, "rows": current_rows})
            current_key = key_value
            current_rows = [row]
        else:
            current_rows.append(row)

    if current_rows:
        groups.append({"group_key": current_key, "rows": current_rows})

    return groups


def build_common_columns(group: Dict[str, Any]) -> tuple[Dict[str, str], List[str]]:
    common_info: Dict[str, str] = {key: "" for key in COMMON_KEYS}
    columns: List[str] = []
    seen_columns: set[str] = set()

    for row in group["rows"]:
        for key in COMMON_KEYS:
            if not common_info[key] and row.get(key):
                common_info[key] = row[key]
        col = row.get("컬럼명")
        if col:
            col_str = str(col)
            if col_str not in seen_columns:
                columns.append(col_str)
                seen_columns.add(col_str)

    return common_info, columns


def build_row_text(group: Dict[str, Any]) -> str:
    common_info, columns = build_common_columns(group)
    payload = {"common": common_info, "columns": columns}
    return json.dumps(payload, ensure_ascii=False, indent=2)


def load_prompt_template() -> str:
    global _PROMPT_CACHE
    if _PROMPT_CACHE is not None:
        return _PROMPT_CACHE
    try:
        with open(DEFAULT_PROMPT_PATH, "r", encoding="utf-8") as f:
            _PROMPT_CACHE = f.read().strip()
            return _PROMPT_CACHE
    except FileNotFoundError:
        _PROMPT_CACHE = (
            "You are given grouped rows from a dataset definition sheet.\n"
            "Generate exactly 8 concise keywords that best represent the dataset, "
            "and a 250-350 character description. Use Korean language. "
            "Return only JSON with keys 'keywords' (array of 8 strings) "
            "and 'description' (string).\n"
            "Row: {row_text}\n"
        )
        return _PROMPT_CACHE


def build_prompt(group: Dict[str, Any], org_name: str | None = None) -> str:
    row_text = build_row_text(group)
    common_info, _ = build_common_columns(group)
    dataset_name = common_info.get("데이터셋명", "")
    template = load_prompt_template()
    prompt = template.replace("{row_text}", row_text)
    prompt = prompt.replace("{기관명}", org_name or "")
    prompt = prompt.replace("{데이터셋명}", dataset_name)
    return prompt


def parse_json_content(content: str) -> Dict[str, Any]:
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        start = content.find("{")
        end = content.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise
        return json.loads(content[start: end + 1])


def _is_gemini_base_url(base_url: str) -> bool:
    return "generativelanguage.googleapis.com" in base_url


def call_openai(prompt: str, base_url: str, api_key: str, model: str) -> Dict[str, Any]:
    if _is_gemini_base_url(base_url):
        url = f"{base_url.rstrip('/')}/models/{model}:generateContent"
        headers = {
            "x-goog-api-key": api_key,
            "Content-Type": "application/json",
        }
        payload = {
            "contents": [{"role": "user", "parts": [{"text": prompt}]}]
        }
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        response.raise_for_status()
        data = response.json()
        print(f"[gemini] model={model}")
        usage = data.get("usageMetadata")
        if usage:
            print(f"[gemini] usageMetadata={json.dumps(usage, ensure_ascii=False)}")
        content = data["candidates"][0]["content"]["parts"][0]["text"]
        return parse_json_content(content)

    url = f"{base_url.rstrip('/')}/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": "You are a precise assistant."},
            {"role": "user", "content": prompt},
        ],
    }
    response = requests.post(url, headers=headers, json=payload, timeout=60)
    response.raise_for_status()
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    return parse_json_content(content)


def generate_summaries(
    rows: List[Dict[str, Any]],
    base_url: str,
    api_key: str,
    model: str,
    include_rows: bool = False,
    sleep_seconds: float = 0.0,
    include_prompt: bool = False,
    org_name: str | None = None,
    progress_callback: Any = None,
) -> List[Dict[str, Any]]:
    if not rows:
        return []

    groups = group_rows(rows)
    total = len(groups)
    results: List[Dict[str, Any]] = []
    if progress_callback:
        progress_callback(0, total, 0)
    for idx, group in enumerate(groups, start=1):
        prompt = build_prompt(group, org_name=org_name)
        result = call_openai(prompt, base_url, api_key, model)
        common_info, columns = build_common_columns(group)
        item: Dict[str, Any] = {
            "row_index": idx,
            "group_key": group.get("group_key"),
            "common": common_info,
            "columns": columns,
            "keywords": result.get("keywords", []),
            "description": result.get("description", ""),
        }
        if include_prompt:
            item["prompt"] = prompt
        if include_rows:
            item["rows"] = group.get("rows")
        results.append(item)
        if progress_callback:
            progress_callback(idx, total, 0)
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)
    return results


def generate_mock_summaries(
    rows: List[Dict[str, Any]],
    include_rows: bool = False,
    include_prompt: bool = False,
    org_name: str | None = None,
) -> List[Dict[str, Any]]:
    if not rows:
        return []

    groups = group_rows(rows)
    results: List[Dict[str, Any]] = []
    for idx, group in enumerate(groups, start=1):
        common_info, columns = build_common_columns(group)
        keywords = columns[:8] if columns else list(common_info.values())[:8]
        item: Dict[str, Any] = {
            "row_index": idx,
            "group_key": group.get("group_key"),
            "common": common_info,
            "columns": columns,
            "keywords": keywords,
            "description": json.dumps({"common": common_info, "columns": columns}, ensure_ascii=False),
        }
        if include_prompt:
            item["prompt"] = build_prompt(group, org_name=org_name)
        if include_rows:
            item["rows"] = group.get("rows")
        results.append(item)
    return results
