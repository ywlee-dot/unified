# -*- coding: utf-8 -*-
import openpyxl
from typing import Dict, List, Any, Optional
import os

class TemplateParser:
    """
    엑셀 템플릿 파일을 파싱하여 설문 구조를 생성합니다.
    """
    
    def __init__(self, template_path: str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
            
        self.template_path = template_path
        self.workbook = openpyxl.load_workbook(template_path, data_only=True)
        self.questions = []
        self.logic_rules = []

    def parse(self) -> Dict[str, Any]:
        """
        전체 템플릿을 파싱합니다.
        """
        self.questions = self._parse_questions()
        self.logic_rules = self._parse_logic_rules()
        
        return {
            "questions": self.questions,
            "logic_rules": self.logic_rules
        }

    def _parse_questions(self) -> List[Dict[str, Any]]:
        """
        '질문' 시트를 파싱합니다.
        컬럼: [1] 섹션, [2] 질문ID, [3] 질문유형, [4] 질문내용, [5] 필수여부, [6] 옵션
        """
        questions = []
        if "질문" not in self.workbook.sheetnames:
            return questions
            
        sheet = self.workbook["질문"]
        
        # 헤더 제외하고 데이터 행 순회
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            section = row[0]
            q_id = row[1]
            if not q_id: continue # 질문 ID가 없으면 중단
            
            q_type = row[2]
            q_text = row[3]
            is_required = True if str(row[4]).upper() == 'Y' else False
            options_raw = row[5]
            
            # 옵션 처리 (콤마 구분)
            options = []
            if options_raw:
                if str(options_raw).strip() == "[DATA_LIST]":
                    options = "[DATA_LIST]" # 특수 플레이스홀더 유지
                else:
                    options = [opt.strip() for opt in str(options_raw).split(",") if opt.strip()]
            
            questions.append({
                "section": str(section) if section else "",
                "question_id": str(q_id),
                "question_type": str(q_type),
                "question_text": str(q_text),
                "required": is_required,
                "options": options
            })
            
        return questions

    def _parse_logic_rules(self) -> List[Dict[str, Any]]:
        """
        '조건부로직' 시트를 파싱합니다.
        컬럼: [1] 로직ID, [2] 조건_질문ID, [3] 조건_응답값, [4] 액션, [5] 대상_질문ID
        """
        logic_rules = []
        if "조건부로직" not in self.workbook.sheetnames:
            return logic_rules
            
        sheet = self.workbook["조건부로직"]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            logic_id = row[0]
            if not logic_id: continue
            
            trigger_qid = row[1]
            trigger_values_raw = row[2]
            action = row[3]
            target_qids_raw = row[4]
            
            # 쉼표로 구분된 값들을 리스트로 변환
            trigger_values = [v.strip() for v in str(trigger_values_raw).split(",") if v.strip()]
            target_qids = [qid.strip() for qid in str(target_qids_raw).split(",") if qid.strip()]
            
            logic_rules.append({
                "logic_id": str(logic_id),
                "trigger_question_id": str(trigger_qid),
                "trigger_values": trigger_values,
                "action": str(action),
                "target_question_ids": target_qids
            })
            
        return logic_rules

    def get_survey_template(self) -> Dict[str, Any]:
        """
        파싱된 템플릿 정보를 반환합니다.
        """
        return {
            "questions": self.questions if self.questions else self._parse_questions(),
            "logic_rules": self.logic_rules if self.logic_rules else self._parse_logic_rules()
        }
